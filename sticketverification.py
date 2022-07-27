## This is for sticket log verification

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
#make patient list configurable
# filter in sticket log
import traceback
import base64
import configparser
import sys
from csv import DictReader
import json
import xlrd
from openpyxl.styles.differential import DifferentialStyle
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, \
    ElementNotInteractableException, ElementClickInterceptedException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import xlwt
#from xlwt import Workbook
import time
import csv
from csv import DictReader
from colorama import Fore, Back, Style
import logging
import os
import shutil
import datetime as x
from datetime import date, datetime,timedelta
import pytz
import re
config = configparser.RawConfigParser()
config.read("locator-config.properties")

def setup(val,downloaddefault):
    if val.lower() == "chrome":
        options = Options()
        options.add_argument("--disable-notifications")
        # change for user
        # userdatadir=config.get("runner","user-data-dir")
        # options.add_argument("user-data-dir={}".format(userdatadir))
        # # options.add_experimental_option("detach", True)
        # self.driver = webdriver.Chrome(executable_path="../chromedriver.exe", chrome_options=options)
        # change for user
        preferences = {
            "download.default_directory":downloaddefault}
        options.add_experimental_option("prefs", preferences)
        # change for user
        driverpath=config.get("runner","chromedriverpath")
        driver = webdriver.Chrome(driverpath,options=options)
        driver.get(config.get("runner","Logout_URL"))
        title = driver.get(config.get("runner","Logout_URL"))
        driver.maximize_window()
    return driver


def action_click(element):
    try:
        element.click()
    except (ElementNotInteractableException, ElementClickInterceptedException):
        driver.execute_script("arguments[0].click();", element)

def login(driver, loc):
    # loc = "C:\\Users\\ssrivastava\\Documents\\credentials.xls"
    workbook = xlrd.open_workbook(loc)
    sheet = workbook.sheet_by_index(0)
    uname = driver.find_element_by_id("edit-name")
    pwd = driver.find_element_by_id("edit-pass")
    uname.send_keys(sheet.cell_value(1, 0))
    pwd.send_keys(sheet.cell_value(1, 1))
    driver.find_element_by_id("edit-submit").click()
    # reason for login
    actions = ActionChains(driver)
    WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
    reason = driver.find_element_by_xpath("//textarea[@id=\"reason_textbox\"]")
    actions.click(reason)
    actions.send_keys_to_element(reason, "https://redmine2.cozeva.com/issues/7662 ")
    actions.perform()
    driver.find_element_by_id("edit-submit").click()


def makedir(foldername):
    path1 = str(foldername)
    if not os.path.exists(path1):
        try:
            os.mkdir(path1)
            return path1
        except OSError as error:
            print(error)
            return False
    else:
        try:
            shutil.rmtree(path1)
            os.mkdir(path1)
            return path1
        except OSError as error:
            print(error)
            return False

def ajax_preloader_wait(driver):
    #time.sleep(1)
    WebDriverWait(driver, 300).until(
        EC.invisibility_of_element((By.XPATH, "//div[contains(@class,'ajax_preloader')]")))
    time.sleep(1)

def date_time():
    today = date.today()
    # print("Today's date:", today)
    tz_In = pytz.timezone('Asia/Kolkata')
    datetime_In = datetime.now(tz_In)
    # print("IN time:", datetime_In.strftime("%I;%M;%S %p"))
    time = datetime_In.strftime("[%I-%M-%S %p]")
    now = str(today) + time
    print(now)
    # logger.info("Date and Time captured!")
    return (now)


#if empty it will return Not found , if fail it will return  : Fail , Please check driver.getURL,#IF Pass ; then just "Pass"
def validate_all_columns(list):
    result = all(element == list[0] for element in list)
    status=""
    if (result):
        status="PASS"
        return status
        #print("All the elements are Equal")
    else:
        status="FAIL"
        return status
        #print("All Elements are not equal")

def validate_date_time_format(s):
    format = "%m/%d/%Y %H:%M:%S"
    status=""
    try:
        x.datetime.strptime(s, format)
        print("This is the correct date string format.")
        status="PASS"
        return status
    except ValueError:
        print("This is the incorrect date string format. It should be YYYY-MM-DD")
        status="Unexpected Value"
        return status




def validate_is_a_proper_string(s):
    status=""
    if(len(s)==0):
        status="Value is missing"
        logger.info(str(s)+" is empty ")
        return status
    if(len(s)>3):
        status="PASS"
        logger.info(str(s)+"is a proper string")
        return status
    if(len(s)<3):
        status="Unexpected Value "
        logger.info(str(s)+" has unexpected value")
        return status

def validate_dob_original(s):
    format = "%m/%d/%Y"
    status = ""
    try:
        x.datetime.strptime(s, format)
        logger.info("This is the correct date string format.")
        status="PASS"
        return status
    except ValueError:
        status="Unexpected Value"
        return status
        logger.info("This is the incorrect date string format. It should be YYYY-MM-DD")


def validate_created(created):
    status = validate_date_time_format(created)
    return status

def validate_last_updated(last_updated):
    status = validate_date_time_format(last_updated)
    return status
def validate_created_by(created_by):
    status = validate_is_a_proper_string(created_by)
    return status
def validate_last_updated_by(last_updated_by):
    status = validate_is_a_proper_string(last_updated_by)
    return status

def validate_patient_status(patient):
    status = validate_is_a_proper_string(patient)
    return status

def validate_dob(dob):
    status = validate_dob_original(dob)
    return status

def validate_member_id_status(member_id):
    status = validate_is_a_proper_string(member_id)
    return status
def validate_member_phone(member_phone):
    status = validate_is_a_proper_string(member_phone)
    return status
def validate_pcp(pcp):
    status = validate_is_a_proper_string(pcp)
    return status
def validate_latest_note(latest_note):
    status = validate_is_a_proper_string(latest_note)
    return status


#
def verify_filter(driver,filter_name,value):
    #click on filter icon
    filter=driver.find_element_by_xpath(config.get("sticket-log-locator","filter_list"))
    action_click(filter)
    if(filter_name=="created"):
        #extract date value
        original_format="%m/%d/%Y %H:%M:%S"
        date_original_value = x.datetime.strptime(value, original_format)
        format = "%m/%d/%Y"
        date_value_string=date_original_value.strftime(format)
        date_value_date = datetime.strptime(date_value_string, format)
        print(date_value_date)

        #send one day less to upper
        yesterday = date_value_date - timedelta(days=1)
        #convert yeaterday to proper format
        yesterday_date_value_string = yesterday.strftime(format)

        created_upper_input=driver.find_element_by_xpath(config.get("sticket-log-locator","created_box_1"))
        created_upper_input.send_keys(yesterday_date_value_string)

        #send current date  to lower

        created_lower_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_box_2"))
        created_lower_input.send_keys(date_value_string)

        #click on apply

        apply=driver.find_element_by_xpath(config.get("sticket-log-locator", "apply_xpath"))
        action_click(apply)
        #check number of returned records

        ajax_preloader_wait(driver)
        num_of_entries = find_number_of_rows(driver)

        #if >1 pass else fail

        if(num_of_entries>=1):
            return "PASS"
        else:
            return "FAIL"

    if(filter_name=="last_updated"):

        created_upper_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_box_1"))
        created_upper_input.clear()
        created_lower_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_box_2"))
        created_lower_input.clear()
        original_format = "%m/%d/%Y %H:%M:%S"
        date_original_value = x.datetime.strptime(value, original_format)
        format = "%m/%d/%Y"
        date_value_string = date_original_value.strftime(format)
        date_value_date = datetime.strptime(date_value_string, format)
        print(date_value_date)

        # send one day less to upper
        yesterday = date_value_date - timedelta(days=1)
        # convert yeaterday to proper format
        yesterday_date_value_string = yesterday.strftime(format)

        last_updated_upper_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "last_updated_box_1"))
        last_updated_upper_input.send_keys(yesterday_date_value_string)

        # send current date  to lower

        last_updated__lower_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "last_updated_box_2"))
        last_updated__lower_input.send_keys(date_value_string)

        # click on apply

        apply = driver.find_element_by_xpath(config.get("sticket-log-locator", "apply_xpath"))
        action_click(apply)
        # check number of returned records

        ajax_preloader_wait(driver)
        num_of_entries = find_number_of_rows(driver)

        # if >1 pass else fail

        if (num_of_entries>= 1):
            return "PASS"
        else:
            return "FAIL"

    if(filter_name=="created_by"):
        created_upper_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_box_1"))
        created_upper_input.clear()
        created_lower_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_box_2"))
        created_lower_input.clear()
        last_updated_upper_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "last_updated_box_1"))
        last_updated_upper_input.clear()
        last_updated__lower_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "last_updated_box_2"))
        last_updated__lower_input.clear()
        created_by=driver.find_element_by_xpath(config.get("sticket-log-locator", "created_by"))
        created_by.send_keys(value)
        # click on apply

        apply = driver.find_element_by_xpath(config.get("sticket-log-locator", "apply_xpath"))
        action_click(apply)
        # check number of returned records

        ajax_preloader_wait(driver)

        num_of_entries = find_number_of_rows(driver)
        # if >1 pass else fail
        print("Number of entries in created by "+str(num_of_entries))
        if (num_of_entries>= 1):
            return "PASS"
        else:
            return "FAIL"

    if(filter_name=="last_updated_by"):
        #print("In llast updatedby ")
        created_upper_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_box_1"))
        created_upper_input.clear()
        created_lower_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_box_2"))
        created_lower_input.clear()
        last_updated_upper_input = driver.find_element_by_xpath(config.get("sticket-log-locator", "last_updated_box_1"))
        last_updated_upper_input.clear()
        last_updated__lower_input = driver.find_element_by_xpath(
            config.get("sticket-log-locator", "last_updated_box_2"))
        last_updated__lower_input.clear()
        created_by = driver.find_element_by_xpath(config.get("sticket-log-locator", "created_by"))
        created_by.clear()

        last_updated_by = driver.find_element_by_xpath(config.get("sticket-log-locator", "last_updated_by"))
        last_updated_by.send_keys(value)

        # click on apply

        apply = driver.find_element_by_xpath(config.get("sticket-log-locator", "apply_xpath"))
        action_click(apply)
        # check number of returned records

        ajax_preloader_wait(driver)
        num_of_entries = find_number_of_rows(driver)
        # if >1 pass else fail

        if (num_of_entries >= 1):
            return "PASS"
        else:
            return "FAIL"

    # if(filter_name=="patient"):











def open_customer_messaging(cust_id):
    sm_customer_id = cust_id  # enter customer_id
    session_var = 'app_id=cozeva_messages&custId=' + str(sm_customer_id) + '&orgId=' + str(sm_customer_id)
    encoded_string = base64.b64encode(session_var.encode('utf-8'))
    driver.get(config.get("runner","URL")+"cozeva_messages?session=" + encoded_string.decode('utf-8')+"&tab=MessageList&label=Inbox&first_load=true")

# def try_to_click(element):
#     attempt=0
#     while(attempt<3):

def find_number_of_rows(driver):
    num_of_rows_total_xpath = config.get("sticket-log-locator", "num_of_rows_total_xpath")
    element = driver.find_elements_by_xpath(num_of_rows_total_xpath)
    num_of_entries = len(element)-1
    return num_of_entries

def find_number_of_columns(driver,column_xpath):
    element = driver.find_elements_by_xpath(column_xpath)
    num_of_columns = len(element)
    print("Number of columns is " +str(num_of_columns))
    return num_of_columns

def extract_name_of_columns(driver,column_xpath):
    element = driver.find_elements_by_xpath(column_xpath)
    num_of_columns = len(element)
    header=[]
    for i in range(1,num_of_columns):
        header=driver.find_element_by_xpath(column_xpath+"["+str(i)+"]")
        header_text=header.get_attribute("innerHTML")
        header.append(header_text)
    return header

def extract_patient_id(href):
    cozeva_id = re.search('/patient_detail/(.*)?session', href)
    return(cozeva_id.group(1).replace("?", ""))

def apply_conditional_formatting(sh1):
    red_text = Font(color="9C0006")
    red_fill = PatternFill(bgColor="FFC7CE")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="FAIL", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']

    green_text = Font(color="00FF00FF")
    green_fill = PatternFill(bgColor="0000FF00")
    dxf = DifferentialStyle(font=green_text, fill=green_fill)
    rule1 = Rule(type="containsText", operator="containsText", text="PASS", dxf=dxf)
    rule1.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']

    sh1.conditional_formatting.add('B1:B10000', rule)
    sh1.conditional_formatting.add('B1:B10000', rule1)

def open_registry_page(customer_id):
    customer_list_url = []
    sm_customer_id = str(customer_id)
    sm_customer_id = sm_customer_id.split(".")[0]
    session_var = 'app_id=registries&custId=' + str(sm_customer_id) + '&payerId=' + str(
        sm_customer_id) + '&orgId=' + str(sm_customer_id)
    encoded_string = base64.b64encode(session_var.encode('utf-8'))
    customer_list_url.append(encoded_string)
    for idx, val in enumerate(customer_list_url):
        driver.get(config.get("runner","URL")+"registries?session=" + val.decode('utf-8'))

def retrieve_modal_attributes(logger,cozeva_id):
    number_of_logs_in_modal=driver.find_elements_by_xpath(config.get("locator","no_of_logs_xpath"))
    logger.info("Number of sticket/contact for "+str(cozeva_id)+ " is "+ str(len(number_of_logs_in_modal)-1))




def validate_time_displayed(logger,time_displayed):
    time_displayed_list = time_displayed.split("<br>")
    print(time_displayed_list)
    timestamp=time_displayed_list
    date_string = time_displayed_list[0]
    format = "%m/%d/%Y"
    date_status=""
    try:
        x.datetime.strptime(date_string, format)
        date_status="PASS"
        print("This is the correct date string format.")
        logger.info("This is the correct date string format.")
    except ValueError:
        date_status="FAIL"
        print("This is the incorrect date string format. It should be MM-DD-YYYY")
        logger.error("This is the incorrect date string format. It should be MM-DD-YYYY")

    time_string = time_displayed_list[1]
    format2 = "%H:%M:%S"
    time_status = ""
    try:
        x.datetime.strptime(date_string, format)
        time_status = "PASS"
        print("This is the correct time string format.")
        logger.info("This is the correct time string format.")
    except ValueError:
        time_status = "FAIL"
        print("This is the incorrect time string format. It should be %H:%M:%S")
        logger.error("This is the incorrect time string format. It should be %H:%M:%S")

    if(date_status=="PASS" and time_status=="PASS"):
        logger.info("Validate Date and Time Displayed -Done - PASS")
        return "PASS"
    else:
        logger.info("Validate Date and Time Displayed -Done - FAIL")
        return "FAIL"



def validate_sender_displayed(logger,sender_displayed):
    num_of_words=len(sender_displayed.split())
    if(num_of_words>1):
        logger.info("Sender Displayed is valid")
        return "PASS"
    else:
        logger.error("Please check Sender name reflection")
        return "FAIL"

def validate_signature_displayed(logger,signature_displayed):
    if "Sent by" in signature_displayed:
        logger.info("Signature Sent By Displayed is valid")
        return "PASS"
    else:
        logger.error("Please check Signature Sent By reflection")
        return "FAIL"
def assert_added(logger):
    time_displayed=driver.find_element_by_xpath(config.get("locator","time_displayed_xpath")).get_attribute("innerHTML")
    logger.info("Time Displayed in sticket is "+str(time_displayed))
    time_displayed_status=validate_time_displayed(logger,time_displayed)
    print("Time Displayed Status ", time_displayed_status)

    sender_displayed = driver.find_element_by_xpath(config.get("locator", "sender_displayed_xpath")).get_attribute(
        "innerHTML")
    logger.info("Sender Displayed in sticket is " + str(sender_displayed))
    sender_displayed_status=validate_sender_displayed(logger,sender_displayed)

    signature_displayed = driver.find_element_by_xpath(config.get("locator", "signature_xpath")).get_attribute(
        "innerHTML")
    logger.info("Signature Displayed in sticket is " + str(signature_displayed))
    signature_displayed_status=validate_signature_displayed(logger,signature_displayed)

    logger.info("Date and Time Test Case "+str(time_displayed_status)+" Sender Displayed Test Case in sticket is "+str(sender_displayed_status)+" Signature Displayed Test Case in sticket is  "+str(signature_displayed_status))
    if(time_displayed_status==sender_displayed_status==signature_displayed_status=="PASS"):
        return "PASS"
    else:
        return "FAIL"

def add_sticket(logger,cozeva_id):
    assert_added_status = "N/A"
    delete_status = "N/A"
    ajax_preloader_wait(driver)
    if len(driver.find_elements_by_xpath(
            config.get("locator", "xpath_patient_Header_Dropdown_Arrow"))) != 0:
        action_click(driver.find_element_by_xpath(config.get("locator", "patient_drop_down")))
        action_click(driver.find_element_by_xpath(config.get("locator", "messages_arrow")))
        action_click(driver.find_element_by_xpath(config.get("locator", "new_sticket")))
        logger.info("clicked on New Sticket for " + str(cozeva_id))
        ajax_preloader_wait(driver)
        retrieve_modal_attributes(logger,cozeva_id)
        driver.find_element_by_xpath(config.get("locator", "sticket_modal")).send_keys(config.get("runner", "text"))
        logger.info("Entered text for sticket ")
        time.sleep(2)
        action_click(driver.find_element_by_xpath(config.get("locator", "save_button")))
        logger.info("Saved sticket")
        time.sleep(5)
        added = 1
        WebDriverWait(driver, 20).until(
            EC.text_to_be_present_in_element((By.XPATH, '(//*[text()="test!@#@##@ 123"])[1]'), "test!@#@##@ 123"))
        assert_added_status = assert_added(logger)
        logger.info("Assert Added status " + str(assert_added_status))
        return assert_added_status

timestamp=[]
def assert_deleted(): #uses text
    sticket_by_text=driver.find_element_by_xpath(config.get("locator","sticket_by_text_xpath"))
    #sticket_by_timestamp_xpath="//div[@class='col s2 message_time' and normalize-space(text()[1])="+"'"+timestamp[0]+"'"+"and normalize-space(text()[2])="+"'"+timestamp[1]+"'"+"]"
    try:
        WebDriverWait(driver, 30).until(
            EC.invisibility_of_element_located((By.XPATH, config.get("locator", "sticket_by_text_xpath"))))
        return "PASS"
    except TimeoutException:
        print("Failed in assert delete")
        return "FAIL"
#
# def verify_add_sticket(driver, workbook, logger, run_from, customer_id):
#     ws1 = workbook.create_sheet("AddedSticket")
#     sh1 = workbook["AddedSticket"]
#
#         return [cozeva_id,status]

def verify_sticket(driver,workbook,logger,run_from,customer_id):
    try:
        ws1 = workbook.create_sheet("Stickets")
        sh1 = workbook["Stickets"]
        if (run_from == "Cozeva Support"):
            # initialize report
            sh1['A1'].value = "Test Case"
            sh1['A1'].font = Font(bold=True, size=13)
            sh1['B1'].value = "Status"
            sh1['B1'].font = Font(bold=True, size=13)
            sh1['C1'].value = "Comments"
            sh1['C1'].font = Font(bold=True, size=13)
            sh1['A2'] = "Sticket page loads in less than 60 sec"
            sh1['A3'] = "All Columns appearing properly "
            sh1['A4'] = "Column Data Display"
            sh1['A5'] = "Created"
            sh1['A6'] = " Last Updated"
            sh1['A7'] = "Created By"
            sh1['A8'] = "Last Updated By"
            sh1['A9'] = "Patient"
            sh1['A10'] = "DOB"
            sh1['A11'] = "Member ID"
            sh1['A12'] = "Member Phone #"
            sh1['A13'] = "PCP"
            sh1['A14'] = "Latest Note"
            sh1['A15'] = "Filter Status"
            sh1['A16'] = "Created"
            sh1['A17'] = "Last Updated"
            sh1['A18'] = "Created By"
            sh1['A19'] = "Last Updated By"
            sh1['A21']= "Reflection of Added sticket"
            open_registry_page(customer_id)
            ajax_preloader_wait(driver)
            logger.info("Opened customer registry" + str(config.get("runner", "customer")))
            # click on filter icon

            filter_icon = driver.find_element_by_xpath(config.get("locator", "filter_list"))
            action_click(filter_icon)

            # sort the registry

            sort_by = driver.find_element_by_xpath(config.get("locator", "sort_by_xpath"))
            action_click(sort_by)

            denominator_option = driver.find_element_by_xpath(config.get("locator", "denominator_option_xpath"))
            action_click(denominator_option)

            apply_button = driver.find_element_by_xpath(config.get("locator", "apply_button_xpath"))
            action_click(apply_button)

            # click on first metric

            first_metric = driver.find_element_by_xpath(config.get("locator", "first_metric_xpath"))
            action_click(first_metric)

            # wait for page to load

            ajax_preloader_wait(driver)
            name_header_xpath = config.get("locator", "table_header_xpath")
            WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, name_header_xpath)))
            # open patients tab

            patient_tab = driver.find_element_by_xpath(config.get("locator", "patient_xpath"))
            action_click(patient_tab)

            # wait for page to load

            ajax_preloader_wait(driver)
            table_header_patient_xpath = config.get("locator", "table_header_patient_xpath")
            WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, table_header_patient_xpath)))
            # store the list of patients

            patient_num = 0
            patient_id_links = driver.find_elements_by_xpath(config.get("locator", "patient_name_list"))
            # click first patient and add sticket
            for patient_num in range(1, 2):
                patient_xpath_final = "(" + config.get("locator", "patient_name_list") + ")" + "[" + str(
                    patient_num) + "]"
                patient_link = driver.find_element_by_xpath(patient_xpath_final)
                cozeva_id = extract_patient_id(patient_link.get_attribute("href"))
                # click on patient
                action_click(patient_link)

            driver.switch_to.window(driver.window_handles[1])
            window_switched = 1
            # returns Pass only if added data is reflected properly in the modal
            status=add_sticket(logger, cozeva_id)
            driver.switch_to.window(driver.window_handles[0])
            # open customer messaging
            open_customer_messaging(customer_id)
            logger.info("Navigating to customer messaging ")
            # ascertain time to load
            time_to_load_start = datetime.now()
            ajax_preloader_wait(driver)
            time_to_load_end = datetime.now()
            time_to_load = time_to_load_end - time_to_load_start
            print("Time to load page", time_to_load)
            logger.info("Time to load messaging page " + str(time_to_load))
            # click on sticket drop down
            sticket_drop_down = driver.find_element_by_xpath(
                config.get("sticket-log-locator", "sticket_dropdown_icon_xpath"))
            action_click(sticket_drop_down)
            logger.info("Clicked on collapsible drop down")
            # scroll down page
            sticket_log_link = driver.find_element_by_xpath(config.get("sticket-log-locator", "sticket_log_xpath"))
            driver.execute_script("arguments[0].scrollIntoView();", sticket_log_link)

            # click on sticket log

            action_click(sticket_log_link)
            time_to_load_sticket_page_start = datetime.now()

            # record time for page load
            created_column_xpath = config.get("sticket-log-locator", "created_column_xpath")
            try:
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, created_column_xpath)))
                page_load_status = "PASS"
            except TimeoutException:
                logger.error("Page is taking more than 60 seconds to load ")
                page_load_status = "FAIL"

            time_to_load_sticket_page_end = datetime.now()

            time_to_load_sticket_page = time_to_load_sticket_page_end - time_to_load_sticket_page_start
            print("Time to load sticket ", time_to_load_sticket_page)
            logger.info("Time to load sticket page " + str(time_to_load_sticket_page))

            # count number of rows

            num_of_rows_total = find_number_of_rows(driver)
            try:
                if (num_of_rows_total >=1):
                    sh1['A20'] = "Number of records"
                    sh1['C20'] = num_of_rows_total
            except:
                sticket_added_reflection = "FAIL"
                raise Exception("Sticket records is empty")

            column_set_string = config.get("runner", "column_set")
            column_set = list(column_set_string.split(","))
            print("column set" + str(column_set))
            column_xpath = "(//tr[@role='row'])[1]//child::th"
            column_set_match = ""
            all_column_status = []
            all_columns_comment = ""
            if (find_number_of_columns(driver, column_xpath) - 1 == len(column_set)):
                column_set_match = "PASS"
                created = []
                last_updated = []
                created_by = []
                last_updated_by = []
                patient = []
                dob = []
                member_id = []
                member_phone = []
                pcp = []
                latest_note = []

                # will check from first four records
                for i in range(2, 3):
                    row_xpath_2 = config.get("sticket-log-locator", "row_xpath_2") + str(i) + "]"

                    created_xpath = row_xpath_2 + "//child::td[1]"
                    last_updated_xpath = row_xpath_2 + "//child::td[2]"
                    created_by_xpath = row_xpath_2 + "//child::td[3]"
                    last_updated_by_xpath = row_xpath_2 + "//child::td[4]"
                    patient_xpath = row_xpath_2 + "//child::td[5]//child::a"
                    dob_xpath = row_xpath_2 + "//child::td[6]"
                    member_id_xpath = row_xpath_2 + "//child::td[7]//div"
                    member_phone_xpath = row_xpath_2 + "//child::td[8]"
                    pcp_xpath = row_xpath_2 + "//child::td[9]"
                    latest_note_xpath = row_xpath_2 + "//child::td[10]//span[@style='display: inline-block;word-break: break-all;']"


                    try:
                        data="created_value"
                        created_text = driver.find_element_by_xpath(created_xpath).get_attribute("innerHTML")
                        created.append(created_text)

                        data="last_updated_value"
                        last_updated_text = driver.find_element_by_xpath(last_updated_xpath).get_attribute("innerHTML")
                        last_updated.append(last_updated_text)

                        data = "last_updated_value"
                        created_by_text = driver.find_element_by_xpath(created_by_xpath).get_attribute("innerHTML")
                        created_by.append(created_by_text)

                        data = "last_updated_by_value"
                        last_updated_by_text = driver.find_element_by_xpath(last_updated_by_xpath).get_attribute(
                            "innerHTML")
                        last_updated_by.append(last_updated_by_text)

                        data = "patient_value"
                        patient_text = extract_patient_id(
                            driver.find_element_by_xpath(patient_xpath).get_attribute("href"))
                        patient.append(patient_text)

                        data = "dob_value"
                        dob_text = driver.find_element_by_xpath(dob_xpath).get_attribute("innerHTML")
                        dob.append(dob_text)

                        data = "member_id_value"
                        try:
                            member_id_text = driver.find_element_by_xpath(member_id_xpath).get_attribute("innerHTML")
                            member_id.append(member_id_text)
                        except NoSuchElementException:
                            member_id.append(" ")
                            pass



                        data = "member_phone_value"
                        try:
                            member_phone_text = driver.find_element_by_xpath(member_phone_xpath).get_attribute(
                                "innerHTML")
                            member_phone.append(member_phone_text)
                        except NoSuchElementException:
                            member_phone.append(" ")
                            pass



                        data = "pcp_value"
                        pcp_text= driver.find_element_by_xpath(pcp_xpath).get_attribute("innerHTML")
                        pcp.append(pcp_text)

                        data = "latest_note_value"
                        latest_note_text = driver.find_element_by_xpath(latest_note_xpath).get_attribute("innerHTML")
                        latest_note.append(latest_note_text)
                    except NoSuchElementException:
                        print(str(data)+"Not found")
                        pass

                    print(created)
                    print(last_updated)
                    print(created_by)
                    print(last_updated_by)
                    print(patient)
                    print(dob)
                    print(member_id)
                    print(member_phone)
                    print(pcp)
                    print(latest_note)

                    all_columns_status = []
                created_status=""
                last_updated_status=""
                created_by_status=""
                last_updated_by_status=""
                patient_status=""
                dob_status=""
                member_id_status=""
                member_phone_status=""
                pcp_status=""
                latest_note_status=""
                for i in range(0, 1):
                    created_status = validate_created(created[i])
                    all_columns_status.append(created_status)
                    logger.info("Validated created_status " + str(created_status))

                    last_updated_status = validate_last_updated(last_updated[i])
                    all_columns_status.append(last_updated_status)
                    logger.info("Validated last_updated_status" + str(last_updated_status))

                    created_by_status = validate_created_by(created_by[i])
                    all_columns_status.append(created_by_status)
                    logger.info("Validated created_by_status" + str(created_by_status))

                    last_updated_by_status = validate_last_updated_by(last_updated_by[i])
                    all_columns_status.append(last_updated_by_status)
                    logger.info("Validated last_updated_by_status" + str(last_updated_by_status))

                    patient_status = validate_patient_status(patient[i])
                    all_columns_status.append(patient_status)
                    if (patient[i] == cozeva_id):
                        print("patient[i]"+str(patient[i]))
                        print("cozevaid" + str(cozeva_id))
                        sticket_added_reflection = "PASS"
                    logger.info("Validated patient_status" + str(patient_status))

                    dob_status = validate_dob(dob[i])
                    all_columns_status.append(dob_status)
                    logger.info("Validated dob_status" + str(dob_status))

                    member_id_status = validate_member_id_status(member_id[i])
                    all_columns_status.append(member_id_status)
                    logger.info("Validated member_id_status" + str(member_id_status))

                    member_phone_status = validate_member_phone(member_phone[i])
                    all_columns_status.append(member_phone_status)
                    logger.info("Validated member_phone_status" + str(member_phone_status))

                    pcp_status = validate_pcp(pcp[i])
                    all_columns_status.append(pcp_status)
                    logger.info("Validated pcp_status" + str(pcp_status))

                    latest_note_status = validate_latest_note(latest_note[i])
                    all_columns_status.append(latest_note_status)

                    logger.info("Validated latest_note_status " + str(latest_note_status))

                all_column_status = validate_all_columns(all_columns_status)
                logger.info("Result of all columns  " + str(all_columns_status))

                # verifies from second record
                random = 0
                created_filter_status=""
                last_updated_filter_status=""
                created_by_filter_status=""
                last_updated_by_filter_status=""
                if (created_status == "PASS"):
                    created_filter_status = verify_filter(driver, "created", created[random])
                    print("Created Filter Status"+str(created_filter_status))
                if (last_updated_status == "PASS"):
                    last_updated_filter_status = verify_filter(driver, "last_updated", last_updated[random])

                if (created_by_status == "PASS"):
                    created_by_filter_status = verify_filter(driver, "created_by", created_by[random])
                    print("Createdby Filter Status" + str(created_by_filter_status))
                #

                if (last_updated_by_status == "PASS"):
                    last_updated_by_filter_status = verify_filter(driver, "last_updated_by",
                                                                  last_updated_by[random])
                    print(last_updated_by_filter_status)

                #delete sticket



                sh1['B2'] = page_load_status
                sh1['C2'] = time_to_load_sticket_page
                sh1['B3'] = column_set_match
                sh1['C3'] = all_columns_comment
                sh1['B4'] = all_column_status
                sh1['B5'] = created_status
                sh1['B6'] = last_updated_status
                sh1['B7'] = created_by_status
                sh1['B8'] = last_updated_by_status
                sh1['B9'] = patient_status
                sh1['B10'] = dob_status
                sh1['B11'] = member_id_status
                sh1['B12'] = member_phone_status
                sh1['C11'] = "For onshore customers blank is expected"
                sh1['C12'] = "For onshore customers blank is expected"
                sh1['B13'] = pcp_status
                sh1['B14'] = latest_note_status
                sh1['B16'] = created_filter_status
                sh1['B17'] = last_updated_filter_status
                sh1['B18'] = created_by_filter_status
                sh1['B19'] = last_updated_by_filter_status
                sh1['B21']=sticket_added_reflection
                sh1['C21']=cozeva_id

                try:
                    driver.switch_to.window(driver.window_handles[1])

                    action_click(
                        driver.find_element_by_xpath(config.get("locator", "sticket_delete_icon_for_concerned_text")))
                    logger.info("Clicked on delete ")
                    print("Clicked on Deleted")
                    deleted = 1
                    time.sleep(2)
                    action_click(driver.find_element_by_xpath(config.get("locator", "confirm_button")))
                    logger.info("Confirmed deletion from confirmation modal")
                    time.sleep(4)
                    delete_status = assert_deleted()
                    driver.close()
                    logger.info("Closed patient's tab " + str(cozeva_id))
                    driver.switch_to.window(driver.window_handles[0])

                except NoSuchElementException:
                    print("Delete from main")
                    delete_status="FAIL"
                    pass


                driver.refresh()
                ajax_preloader_wait(driver)
                # record time for page load
                created_column_xpath = config.get("sticket-log-locator", "created_column_xpath")
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, created_column_xpath)))


                patient_text2 = extract_patient_id(
                    driver.find_element_by_xpath(patient_xpath).get_attribute("href"))
                if (patient_text2 == cozeva_id):
                    deleted_status = "PASS"
                else:
                    deleted_status = "FAIL"

                if(deleted_status!="PASS"):
                    sh1.append(("Sticket Delete functionality","PASS"))
                else:
                    sh1.append(("Sticket Delete functionality", "FAIL","Manual intervention required for cozeva id given "))

                apply_conditional_formatting(sh1)
                workbook.save("Report_Sticket_Log.xlsx")

                # sh1['D4'] = "Number of Pages " + str(page)


            else:
                column_set_match = False
                list_of_columns = extract_name_of_columns(driver, column_xpath)
                if (len(list_of_columns) < len([column_set])):
                    column_missing = list(set(column_set).difference(list_of_columns))
                    all_columns_comment = " ".join(column_missing) + " is missing"
                else:
                    all_columns_comment = "extra columns found"
    except Exception as e:
        sh1['A1']="Sticket page empty or failed to load "
        workbook.save("Report_Sticket_Log.xlsx")
        print("Failed to continue sticket " + str(e))



      # try:
        #   count number of records
        #

#create Folder or working directory
dateandtime = date_time()
master_directory=config.get("runner","report_directory_input")
os.chdir(master_directory)
path = makedir(dateandtime)
LOG_FORMAT = "%(levelname)s %(asctime)s - %(message)s"
logging.basicConfig(filename=path + "\\" + "Sticket-Log.log", level=logging.INFO, format=LOG_FORMAT, filemode='w')
logger = logging.getLogger()
#logger.setLevel(logging.INFO)
os.chdir(path)

downloaddefault=config.get("runner","downloaddefault")
makedir(downloaddefault)
driver = setup("Chrome",downloaddefault)
begin_time = datetime.now()
loc = config.get("runner","login_file")

#login
login(driver, loc)
logger.info("Login successful")

#Initialize Worksheet

wb=openpyxl.Workbook()


customer_id="4500"
#def verify_sticket_functionality(driver, workbook, logger, run_from, customer_id):
    # added_status = [cozevaid, status] = verify_add_sticket(driver, workbook, logger, run_from, customer_id)  # worksheetappend
    #cozeva_id=
    # delete_status = verify_delete_sticket(added_status, driver, workbook, logger, run_from, customer_id)  # worksheetappend
    # if (delete_status == False):
    #     print("Failed to delete Sticket " + " for " + str(added_status[0]) + " Customer " + str(
    #         customer_id))  # add formatting if possible

#
# verify_sticket_functionality(driver, wb, logger, "CS", customer_id)

#verify_sticket(driver, workbook, logger, run_from, customer_id)
verify_sticket(driver, wb, logger, "CozevaSupport", customer_id)

driver.quit()








#
#
#
# if(find_number_of_columns(driver,column_xpath)-1==len(column_set)):
#
#
# else:
#
#



# wb2 = load_workbook('template.xlsx')
# wb2.create_sheet('Sticket Log')
# wb2.save('template.xlsx')
# sh1=wb.active
# sh1.title="Sticket Log"
# else:
#     column_set=['Created','Last Updated','Created By','Last Updated By','Patient','DOB','Member ID','Member Phone #','PCP','Latest Note']
#     column_xpath="(//tr[@role='row'])[1]//child::th"
#     if(find_number_of_columns(driver,column_xpath)-1==len(column_set)):
#         created=[]
#         last_updated=[]
#         created_by=[]
#         last_updated_by=[]
#         patient=[]
#         dob=[]
#         member_id=[]
#         member_phone=[]
#         pcp=[]
#         latest_note=[]
#         for i in range(2, 5):
#             row_xpath_2 = config.get("sticket-log-locator", "row_xpath_2") + str(i) + "]"
#
#             created_xpath = row_xpath_2 + "//child::td[1]"
#             last_updated_xpath = row_xpath_2 + "//child::td[2]"
#             created_by_xpath = row_xpath_2 + "//child::td[3]"
#             last_updated_by_xpath = row_xpath_2 + "//child::td[4]"
#             patient_xpath = row_xpath_2 + "//child::td[5]//child::a"
#             dob_xpath = row_xpath_2 + "//child::td[6]"
#             member_id_xpath = row_xpath_2 + "//child::td[7]//div"
#             member_phone_xpath = row_xpath_2 + "//child::td[8]"
#             pcp_xpath = row_xpath_2 + "//child::td[9]"
#             latest_note_xpath = row_xpath_2 + "//child::td[10]//span[@style='display: inline-block;word-break: break-all;']"
#
#             created_text=driver.find_element_by_xpath(created_xpath).get_attribute("innerHTML")
#             created.append(created_text)
#
#             last_updated_text=driver.find_element_by_xpath(last_updated_xpath).get_attribute("innerHTML")
#             last_updated.append(last_updated_text)
#
#             created_by_text=driver.find_element_by_xpath(created_by_xpath).get_attribute("innerHTML")
#             created_by.append(created_by_text)
#
#             last_updated_by_text=driver.find_element_by_xpath(last_updated_by_xpath).get_attribute("innerHTML")
#             last_updated_by.append(last_updated_by_text)
#
#             patient_text=extract_patient_id(driver.find_element_by_xpath(patient_xpath).get_attribute("href"))
#             patient.append(patient_text)
#
#             dob_text=driver.find_element_by_xpath(dob_xpath).get_attribute("innerHTML")
#             dob.append(dob_text)
#
#             member_id_text=driver.find_element_by_xpath(member_id_xpath).get_attribute("innerHTML")
#             member_id.append(member_id_text)
#
#             member_phone_text=driver.find_element_by_xpath(member_phone_xpath).get_attribute("innerHTML")
#             member_phone.append(member_phone_text)
#
#             pcp_text=driver.find_element_by_xpath(pcp_xpath).get_attribute("innerHTML")
#             pcp.append(pcp_text)
#
#             latest_note_text=driver.find_element_by_xpath(latest_note_xpath).get_attribute("innerHTML")
#             latest_note.append(latest_note_text)
#     else:
#         list_of_columns=extract_name_of_columns(driver,column_xpath)
#         if(len(list_of_columns)<len([column_set])):
#             column_missing=list(set(column_set).difference(list_of_columns))
#             status=" ".join(column_missing)+" is missing"
#         else:
#             status="extra columns found "
#
#     print(created)
#     print(last_updated )
#     print(created_by)
#     print(last_updated_by)
#     print(patient)
#     print(dob )
#     print(member_id )
#     print(member_phone)
#     print(pcp)
#     print(latest_note)
#
#


#count total number of entries code

# till Next button is disabled - count entries in a page - scroll down to next -click on nexxt -wait for page to load

# next_button_xpath=config.get("sticket-log-locator","next_button_xpath")
# row_xpath=config.get("sticket-log-locator","row_xpath")
# num_of_entries=0
# page=1
#
# if(config.get("sticket-log-locator","check_number_of_entries")=="False"):
#     num_of_entries="Not checked "
#     page="Not checked "
#
# # x=5
#
#
# if(config.get("sticket-log-locator","check_number_of_entries")=="True"):
#     while (1):
#
#         # if(x==6):
#         #     break
#         # x = x + 1
#         WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, created_column_xpath)))
#         try:
#             WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, next_button_xpath)))
#         except TimeoutException:
#             if(page>1):
#                 pass
#             else:
#                 element = driver.find_elements_by_xpath(row_xpath)
#                 num_of_entries = find_number_of_rows(driver,row_xpath)
#                 logger.info("Number of total entries till page " + str(page) + " is " + str(num_of_entries))
#                 break
#
#
#         time.sleep(3)
#         num_of_entries=num_of_entries+find_number_of_rows(driver,row_xpath)
#         logger.info("Number of total entries till page " + str(page) + " is " + str(num_of_entries))
#         try:
#             WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, next_button_xpath)))
#         except TimeoutException:
#             logger.info("Last Page ")
#             break;
#         next_button = driver.find_element_by_xpath(next_button_xpath)
#         driver.execute_script("arguments[0].scrollIntoView();", next_button)
#         if (next_button.get_attribute("disabled") == "disabled"):
#             break
#         action_click(next_button)
#         page = page + 1
#         ajax_preloader_wait(driver)
#         WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, created_column_xpath)))
#
#     total_time_taken_end = datetime.now()
#     total_time_taken = total_time_taken_end - time_to_load_start
#     print("Total number of entries =", num_of_entries)
#     print("Total number of pages = ", page)
#     print("Total time taken ", total_time_taken)
#     logger.info("Total number of entries =" + str(num_of_entries))
#     logger.info("Total number of pages = " + str(page))
#     logger.info("Total time taken " + str(total_time_taken))
