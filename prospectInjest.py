import base64
import pytz
from openpyxl import Workbook, load_workbook
import os
import sys
from csv import DictReader
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import json
import xlrd
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, ElementNotInteractableException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import xlwt
from xlwt import Workbook
import time
import csv
from csv import DictReader
from colorama import Fore, Back, Style
import logging
import os
import shutil
import datetime as x
from datetime import date, datetime,timedelta
import configparser
import openpyxl
config = configparser.RawConfigParser()
config.read("locator-config.properties")

#generate separate report for errors- if string is not formed completely then throw error and store it in error log


# wb = Workbook
# codingsheet = wb.add_sheet("CodingSheet")

begin = time.time()


def date_time():
    today = date.today()
    # print("Today's date:", today)
    tz_In = pytz.timezone('Asia/Kolkata')
    datetime_In = datetime.now(tz_In)
    # print("IN time:", datetime_In.strftime("%I;%M;%S %p"))
    time = datetime_In.strftime("[%I-%M-%S %p]")
    now = str(today) + time
    print(now)
    # logger.info("Date and Time captured!")
    return (now)


# sys.stdout=open("codingsheet.txt","w")

def makedir(foldername):
    path1 = str(foldername)
    if not os.path.exists(path1):
        try:
            os.mkdir(path1)
            return path1
        except OSError as error:
            print(error)
            return False
    else:
        try:
            shutil.rmtree(path1)
            os.mkdir(path1)
            return path1
        except OSError as error:
            print(error)
            return False

def setup(val):
    # #val.lower() == "firefox":
    # driver = webdriver.Firefox(executable_path=r"C:\\Users\\ssrivastava\\PycharmProjects\\Python_Practise\\driver2\\geckodriver.exe")
    # driver.implicitly_wait(10)
    # title = driver.get("https://stage.cozeva.com/user/login")
    # driver.maximize_window()
    if val.lower() == "chrome":
        options = Options()
        options.add_argument("--disable-notifications")
        # change for user
        options.add_argument("user-data-dir=C:\\Users\\ssrivastava\\PycharmProjects\\codingtooldata1")
        # options.add_experimental_option("detach", True)
        # self.driver = webdriver.Chrome(executable_path="../chromedriver.exe", chrome_options=options)
        # change for user
        preferences = {
            "download.default_directory": "C:\\Users\\ssrivastava\\PycharmProjects\\codingtooldata"}
        options.add_experimental_option("prefs", preferences)
        # change for user
        driverpath = "C:\\Users\\ssrivastava\\PycharmProjects\\Python_Practise\\driver2\\chromedriver.exe"
        driver = webdriver.Chrome(driverpath, options=options)
        driver.get("https://cert.cozeva.com/user/logout")
        title = driver.get("https://cert.cozeva.com/user/login")
        driver.maximize_window()
    return driver


def login(driver):
    loc = config.get("Quality","login_file")
    workbook = xlrd.open_workbook(loc)
    sheet = workbook.sheet_by_index(0)
    uname = driver.find_element_by_id("edit-name")
    pwd = driver.find_element_by_id("edit-pass")
    uname.send_keys(sheet.cell_value(1, 0))
    pwd.send_keys(sheet.cell_value(1, 1))
    driver.find_element_by_id("edit-submit").click()
    # reason for login
    actions = ActionChains(driver)
    reason = driver.find_element_by_xpath("//textarea[@id=\"reason_textbox\"]")
    actions.click(reason)
    actions.send_keys_to_element(reason, "RM7122")
    actions.perform()
    driver.find_element_by_id("edit-submit").click()


def wait_to_load(element_xpath):
    try:
        WebDriverWait(driver, 300).until(EC.visibility_of_element_located((By.XPATH, element_xpath)))
    except TimeoutException:
        print("Time out loaded ")
        pass

def action_click(element):
    webdriver.ActionChains(driver).move_to_element(element).click(element).perform()


def check_exists_by_xpath(driver, xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True


def extractchoiceinfo(taskid, metricid):
    choice_option_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//following-sibling::tr[@metric-id=\"{}\"]//child::td[@class=\"choice_option\"]//div[@class=\"select-wrapper qc_enabled tooltipped\"]".format(
        taskid, metricid)
    choice_options = driver.find_elements_by_xpath(choice_option_xpath)
    choice_value = []
    for choice_option in choice_options:
        choice_value.append(choice_option.get_attribute("data-tooltip"))
    return choice_value


def extractcodeinfo(taskid, metricid,code_entered_xpath):
    metric_name_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//following::div[@class=\"met-name mrb grey-text text-darken-4\"]/span[1]".format(
        taskid)
    code_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//following-sibling::tr[@metric-id={}]//child::td[@class=\"codes\"]//div[@class=\"tooltipped qc_enabled\"]//input".format(
        taskid, metricid)
    bp_code_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//following-sibling::tr[@metric-id={}]//child::td[@class=\"codes\"]//div[@class=\"codebox_inbox\"]//input".format(
        taskid, metricid)
    code_value = []
    metric_name = driver.find_element_by_xpath(metric_name_xpath).text
    if "BP Control" in metric_name or "Body Mass" in metric_name:
        codes = driver.find_elements_by_xpath(bp_code_xpath)
        for mcode in codes:
            code_value.append(mcode.get_attribute('value'))
    else:
        codes = driver.find_elements_by_xpath(code_entered_xpath)
        for mcode in codes:  # mcode is  the code for list of codes obtained
            code_value.append(mcode.get_attribute("data-tooltip"))

    return code_value


def extractdateinfo(taskid, metricid):
    diff = 0
    date_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//following-sibling::tr[@metric-id={}]//child::td[@class=\"service_date\"]//child::input".format(
        taskid, metricid)
    date_elements = driver.find_elements_by_xpath(date_xpath)
    if len(date_elements) == 0:
        diff = 1
        date_xpath = "//tr[@class='metric-tr qc_enabled' and @data-subtask-id={}]//following-sibling::tr[@metric-id={}]//child::td[@class='service_date custom-disabled']//child::input".format(
            taskid, metricid)
        date_elements = driver.find_elements_by_xpath(date_xpath)
    dates = []
    for date_element in date_elements:
        dates.append(date_element.get_attribute("value"))
    return dates


def extractrenderingproviderinfo(taskid, metricid):
    rprov_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//following-sibling::tr[@metric-id={}]//child::td[@class=\"rendering_provider\"]//child::input".format(
        taskid, metricid)
    rprov_elements = driver.find_elements_by_xpath(rprov_xpath)
    if (len(rprov_elements) == 0):
        rprov_xpath = "//tr[@class='metric-tr qc_enabled' and @data-subtask-id={}]//following-sibling::tr[@metric-id={}]//child::td[@class='rendering_provider custom-disabled']//child::input".format(
            taskid, metricid)
        rprov_elements = driver.find_elements_by_xpath(rprov_xpath)
    rendering_provider = []
    for rprov_element in rprov_elements:
        rendering_provider.append(rprov_element.get_attribute("value"))
    return rendering_provider


def extractchoiceinfo_old():
    choice_xpath = "//tr[@class=\"choice_tr saved_data  custom-disabled\"]//child::td[@class=\"choice_option\"]//input"
    choice_options = driver.find_elements_by_xpath(choice_xpath)
    choice_value = []
    for choice_option in choice_options:
        choice_value.append(choice_option.get_attribute("value"))
    return choice_value
def code_to_rows(t, measure_abbreviation, dates, choicevalues, codevalues, ren_provider):
    coded_string = []
    number_of_rows = len(choicevalues)
    list1=[]
    list2=[]
    lst=[list1]
    count=0
    if ("BP" in measure_abbreviation):
        end = int(len(codevalues) / 4)
        # print(end)
        p = 0
        for j in range(0, end):
            lab_value = codevalues[p] + "/" + codevalues[p + 1]
            # print(lab_value)
            office_value = codevalues[p + 2]
            # print(office_value)
            if len(dates) == 0:
                dategiven = "None"
            else:
                try:
                    dategiven = dates[j]
                except IndexError:
                    dategiven = dates[0]
            if len(ren_provider) == 0:
                ren_providergiven = "None"
            else:
                try:
                    ren_providergiven = ren_provider[j]
                except IndexError:
                    ren_providergiven = ren_provider[0]

            s1 = "{}`{}`{}`{}`{}".format(t, measure_abbreviation, dategiven, lab_value, ren_providergiven)
            coded_string.append(s1)
            list1.append(t)
            list1.append(measure_abbreviation)
            list1.append(dategiven)
            list1.append(choicevalues[0])
            list1.append(lab_value)
            list1.append(ren_provider[0])
            s2 = "{}`{}`{}`{}`{}".format(t, measure_abbreviation, dategiven, office_value, ren_providergiven)
            coded_string.append(s2)
            list2.append(t)
            list2.append(measure_abbreviation)
            list2.append(dategiven)
            list2.append(choicevalues[0])
            list2.append(office_value)
            list2.append(ren_provider[0])
            p = p + 4
            lst=[list1,list2]
            count=2
    else:
        for i in range(0, number_of_rows):
            try:
                code = codevalues[i] + "(" + codevalues[i + 1] + ")"
            except:
                code = codevalues[i]

            if len(dates) == 0:
                dategiven = "None"
            else:
                try:
                    dategiven = dates[i]
                except IndexError:
                    dategiven = dates[0]

            if len(ren_provider) == 0:
                ren_providergiven = "None"
            else:
                try:
                    ren_providergiven = ren_provider[i]
                except IndexError:
                    ren_providergiven = ren_provider[0]

            s = "{}`{}`{}`{}`{}".format(t, measure_abbreviation, dategiven, code, ren_providergiven)
            list1.append(t)
            list1.append(measure_abbreviation)
            list1.append(dategiven)
            list1.append(choicevalues)
            list1.append(code)
            list1.append(ren_provider)
            lst=[list1]
            count=1
    return count,lst

def code_to_string(t, measure_abbreviation, dates, choicevalues, codevalues, ren_provider):
    coded_string = []
    number_of_rows = len(choicevalues)
    list1=[]
    list2=[]
    if ("BP" in measure_abbreviation):
        end = int(len(codevalues) / 4)
        # print(end)
        p = 0
        for j in range(0, end):
            lab_value = codevalues[p] + "/" + codevalues[p + 1]
            # print(lab_value)
            office_value = codevalues[p + 2]
            # print(office_value)
            if len(dates) == 0:
                dategiven = "None"
            else:
                try:
                    dategiven = dates[j]
                except IndexError:
                    dategiven = dates[0]
            if len(ren_provider) == 0:
                ren_providergiven = "None"
            else:
                try:
                    ren_providergiven = ren_provider[j]
                except IndexError:
                    ren_providergiven = ren_provider[0]

            s1 = "{}`{}`{}`{}`{}".format(t, measure_abbreviation, dategiven, lab_value, ren_providergiven)
            coded_string.append(s1)
            s2 = "{}`{}`{}`{}`{}".format(t, measure_abbreviation, dategiven, office_value, ren_providergiven)
            coded_string.append(s2)
            p = p + 4
    else:
        for i in range(0, number_of_rows):
            try:
                code = codevalues[i] + "(" + codevalues[i + 1] + ")"
            except:
                code = codevalues[i]

            if len(dates) == 0:
                dategiven = "None"
            else:
                try:
                    dategiven = dates[i]
                except IndexError:
                    dategiven = dates[0]

            if len(ren_provider) == 0:
                ren_providergiven = "None"
            else:
                try:
                    ren_providergiven = ren_provider[i]
                except IndexError:
                    ren_providergiven = ren_provider[0]

            s = "{}`{}`{}`{}`{}".format(t, measure_abbreviation, dategiven, code, ren_providergiven)
            coded_string.append(s)
    return coded_string


def extractdateinfo_old():
    diff = 0
    date_xpath = "//tr[@class=\"choice_tr saved_data  custom-disabled\"]//child::td[@class=\"service_date\"]//descendant::input"

    date_elements = driver.find_elements_by_xpath(date_xpath)
    if len(date_elements) == 0:
        diff = 1
        date_xpath = "//tr[@class='choice_tr saved_data  custom-disabled']//child::td[@class='service_date custom-disabled']//descendant::input"
        date_elements = driver.find_elements_by_xpath(date_xpath)
    dates = []
    for date_element in date_elements:
        dates.append(date_element.get_attribute("value"))
    return dates


def extractrenderingproviderinfo_old():
    rprov_xpath = "//tr[@class=\"choice_tr saved_data  custom-disabled\"]//child::td[@class=\"rendering_provider\"]//descendant::input"
    rprov_elements = driver.find_elements_by_xpath(rprov_xpath)
    if (len(rprov_elements) == 0):
        rprov_xpath = "//tr[@class='choice_tr saved_data  custom-disabled']//child::td[@class='rendering_provider custom-disabled']//descendant::input"
        rprov_elements = driver.find_elements_by_xpath(rprov_xpath)
    rendering_provider = []
    for rprov_element in rprov_elements:
        rendering_provider.append(rprov_element.get_attribute("value"))
    return rendering_provider


def extractcodeinfo_old():
    code_xpath = "//tr[@class=\"choice_tr saved_data  custom-disabled\"]//child::td[@class=\"codes\"]//descendant::input"
    code_values = driver.find_elements_by_xpath(code_xpath)
    code_list = []
    for code_value in code_values:
        code_list.append(code_value.get_attribute("value"))
    return code_list


def report_write_func(report,row,task_id,measure_abb,service_date,choice,code,rendering_provider):
    report.write(row, 0, task_id)
    report.write(row, 1, measure_abb)
    report.write(row, 2, service_date)
    report.write(row, 3, choice)
    report.write(row, 4, code)
    report.write(row,5,rendering_provider)


def verify_codingsheetQuality(driver,workbook,logger,run_from):

    try:
        ws1 = workbook.create_sheet("CodingToolQuality")
        sh1 = workbook["CodingToolQuality"]
        if (run_from == "CozevaSupport"):
            # initialize report
            sh1['A1'].value = "Task"
            sh1['A1'].font = Font(bold=True, size=13)
            sh1['B1'].value = "Choice"
            sh1['B1'].font = Font(bold=True, size=13)
            sh1['C1'].value = "Code"
            sh1['C1'].font = Font(bold=True, size=13)
            sh1['D1'].value = "Dates"
            sh1['D1'].font = Font(bold=True, size=13)
            sh1['E1'].value = "Rendering Provider "
            sh1['E1'].font = Font(bold=True, size=13)
            sh1['F1'].value = "URL"
            sh1['F1'].font = Font(bold=True, size=13)

    except Exception as e:
        print("Report Sheet not created ")
    column_xpath = "//*[@id=\"chart_chase\"]/thead/tr/th[2]"
    wait_to_load(column_xpath)
    wait_to_load(column_xpath)
    hamburger_xpath = "//i[@class=\"material-icons\" and text()=\"menu\"]"
    hamburger = driver.find_element_by_xpath(hamburger_xpath)
    action_click(hamburger)

    side_bar_nav_xpath = "//ul[@id=\"sidenav_slide_out\"]"
    wait_to_load(side_bar_nav_xpath)
    supp_data_xpath = "//i[@class=\"material-icons sidenav_main\"]//following-sibling::span[text()=\"Supplemental Data\"]"
    supp_data = driver.find_element_by_xpath(supp_data_xpath)
    action_click(supp_data)

    wait_to_load(column_xpath)
    # remove other tag

    # other_tag_close_xpath='//div[@class="dt_tag_wrapper"]//child::span[text()="close"]'
    # other_tag_close=driver.find_element_by_xpath(other_tag_close_xpath)
    # action_click(other_tag_close)
    # WebDriverWait(driver, 30).until(EC.invisibility_of_element((By.XPATH, "//div[@class=\"ajax_preloader\"]")))
    # wait_to_load(column_xpath)
    # click on task -> open in new tab ->print element -> Close the tab ->Click on next task
    # random_page_xpath = "//button[text()=\"4\"]"
    # random_page = driver.find_element_by_xpath(random_page_xpath)
    # action_click(random_page)
    filter_list_xpath = "//i[text()=\"filter_list\"]"
    filter_list = driver.find_element_by_xpath(filter_list_xpath)
    filter_list.click()

    #take date from locator config
    new_creation_date_filter_from_xpath = "//input[@name='chart_chase_uploaded_from']"
    new_creation_date_filter_to_xpath = "//input[@name='chart_chase_uploaded_to']"

    date_filter_from = driver.find_element_by_xpath(new_creation_date_filter_from_xpath)
    date_filter_from.clear()
    date_filter_from_input=config.get("Quality","date_from")
    date_filter_from.send_keys(date_filter_from_input)  # start date

    date_filter_to = driver.find_element_by_xpath(new_creation_date_filter_to_xpath)
    date_filter_to.clear()
    date_filter_to_input = config.get("Quality", "date_to")
    date_filter_to.send_keys(date_filter_to_input)  # end date

    # apply link status

    # link_dropdown_xpath='//div[text()="Link Status:"]//parent::div//following-sibling::div'
    # link_dropdown=driver.find_element_by_xpath(link_dropdown_xpath)
    # action_click(link_dropdown)
    #
    # selected_value_from_link_xpath='//div[text()="Link Status:"]//parent::div//following-sibling::div//ul//child::span[text()="Linked"]'
    # select_value=driver.find_element_by_xpath(selected_value_from_link_xpath)
    # action_click(select_value)

    apply_xpath = "//a[text()=\"Apply\"]"
    apply = driver.find_element_by_xpath(apply_xpath)
    apply.click()

    wait_to_load(column_xpath)

    # find no of pages
    WebDriverWait(driver, 100).until(EC.invisibility_of_element((By.XPATH, "//div[@class=\"ajax_preloader\"]")))

    # find no of pages
    pages_xpath = "//button[@class='mdl-button  mdl-button--raised mdl-button--colored']"
    pages = driver.find_elements_by_xpath(pages_xpath)
    print(len(pages))
    num_pages = pages[len(pages) - 1].text


    i = 2
    # for i in range(num_pages+1):
    #     pages_xpath="//button[@class=\"mdl-button \" and text()="{}"]".format(i)

    #########task search from filter ############

    taskid_xpath = "//input[@name='chart_chase_task']"
    parent_taskid_xpath = "//input[@name='chart_chase_parent_task']"
    taskid_filter = driver.find_element_by_xpath(taskid_xpath)
    parent_taskid_filter = driver.find_element_by_xpath(parent_taskid_xpath)

    # taskid_filter.send_keys("Enter string to be searched")
    # parent_taskid_filter.send_keys("Enter key to be searched")

    if(len(pages)>1):
        pages_xpath = "//button[@class='mdl-button ']"
        pages = driver.find_elements_by_xpath(pages_xpath)
        last_page = pages[len(pages) - 1].text
    else:
        last_page=1
    print("Number of pages ", last_page)

    current_page_xpath = "//button[@class='mdl-button  mdl-button--raised mdl-button--colored']"
    current_page = driver.find_element_by_xpath(current_page_xpath).text

    print(current_page)

    next_page_value = int(current_page) + 1

    row = 1

    for i in range(int(current_page), int(last_page) + 1):
        print("Page no", i)
        for x in range(1, 21):
            err = False
            print("\n")
            WebDriverWait(driver, 100).until(
                EC.invisibility_of_element((By.XPATH, "//div[@class=\"ajax_preloader\"]")))
            task_xpath = "//*[@id=\"chart_chase\"]/tbody/tr[%d]/td[2]/div/div/div/a" % x;
            # print(task_xpath)
            measure_abbrev_xpath = "//*[@id=\"chart_chase\"]/tbody/tr[{}]/td[@class=\" chart_chase_service_date\"]//div[@class=\"col s3 enc_meas tooltipped tooltip_init-processed\"]".format(
                x)
            measure_abbrev = driver.find_elements_by_xpath(measure_abbrev_xpath)
            measure_list = []
            if (len(measure_abbrev) > 1):
                for m in measure_abbrev:
                    measure_list.append(m.text)
                measure_abbreviation = measure_list[0]
                print(measure_list)
            else:
                if (len(measure_abbrev) == 0):
                    measure_abbreviation = "Blank"
                else:
                    try:

                        measure_abbreviation = measure_abbrev[0].text
                    except NoSuchElementException:
                        measure_abbreviation = "Blank"

            try:
                task = WebDriverWait(driver, 10).until(
                    EC.invisibility_of_element((By.XPATH, "//div[@class=\"ajax_preloader\"]")))
                task = driver.find_element_by_xpath(task_xpath)
                action_click(task)
            except NoSuchElementException:
                if (i == int(last_page)):
                    break;
                else:
                    print("Error: Task not found")
            t = task.text

            # print(driver.window_handles)
            # print(driver.current_window_handle)
            driver.switch_to.window(driver.window_handles[1])
            coding_table_xpath = "//span[text()=\"Measure\"]"
            wait_to_load(coding_table_xpath)
            metric_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id=\"{}\"]".format(t);

            # check if old task or not - extract codes - match with chart list

            try:
                task_id_left_xpath = "(//div[@class=\"task_labels hide lfloat\"])[1]"
                task_id = driver.find_element_by_xpath(task_id_left_xpath)
            except NoSuchElementException:
                task_id_left_xpath = "//div[@class=\"task_labels label-processed\"]"
                task_id = driver.find_element_by_xpath(task_id_left_xpath)
            task_id_left = task_id.get_attribute("data-task-id")
            old_supp_data = False

            if check_exists_by_xpath(driver, metric_xpath):
                a = driver.find_element_by_xpath(metric_xpath)
                mid = a.get_attribute("metric-id")
                code_entered_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id=\"{}\"]//following-sibling::tr[@metric-id=\"{}\"]/child::td[@class=\"codes\"]/div/div/div[1]".format(
                    t, mid)
                if check_exists_by_xpath(driver, code_entered_xpath) == True:
                    code_entered = driver.find_element_by_xpath(code_entered_xpath)
                    code = code_entered.get_attribute("data-tooltip")
                    metric_abbreviation_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//div[@class=\"quality-metric\"]".format(
                        t)
                    metric_abbreviation_string = driver.find_element_by_xpath(metric_abbreviation_xpath).get_attribute(
                        "data-rch")
                    bad_chars = ["[", "]"]
                    # metric_abbreviation_json = ''.join(i for i in metric_abbreviation_string if not i in bad_chars)
                    # metric_abbreviation_json2=metric_abbreviation_json.split(",{")
                    # print(metric_abbreviation_json2[0])
                    #
                    # metric_abbreviation = json.loads(metric_abbreviation_json2[0])["abbreviation"]
                    metric_name_xpath = "//tr[@class=\"metric-tr qc_enabled\" and @data-subtask-id={}]//following::div[@class=\"met-name mrb grey-text text-darken-4\"]/span[1]".format(
                        t)
                    metric_name = driver.find_element_by_xpath(metric_name_xpath).text
                    choicevalues = extractchoiceinfo(t, mid)
                    codevalues = extractcodeinfo(t, mid,code_entered_xpath)
                    dates = extractdateinfo(t, mid)
                    ren_provider = extractrenderingproviderinfo(t, mid)
                    print(t, " task : entered ")
                    print("Choice : ", choicevalues)
                    print("Code : ", codevalues)
                    print("dates : ", dates)
                    print("Rendering Provider : ", ren_provider)
                    sh1.append((str(t),str(choicevalues),str(codevalues),str(dates),str(ren_provider),str(driver.current_url)))
                    workbook.save("Report_QualityCoding.xlsx")
                    codestring = code_to_string(t, measure_abbreviation, dates, choicevalues, codevalues, ren_provider)
                    print(codestring)
                    # r,rows_list=code_to_rows(t, measure_abbreviation, dates, choicevalues, codevalues, ren_provider)
                    # print(rows_list)
                    sr = 0
                    # print(r)
                    # print(rows_list[1][1])
                    # while sr<r:
                    #     report_write_func(report_sheet,row, rows_list[sr][0],rows_list[sr][1],rows_list[sr][2],rows_list[sr][3],rows_list[sr][4],rows_list[sr][5])
                    #     row=row+1
                    #     report.save(dest_filename)
                    #     sr=sr+1

                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                else:
                    print(t, "Error code not found")
                    get_url = driver.current_url
                    print(t, get_url)
                    sh1.append((str(t), "Error code not found"))
                    row = row + 1
                    workbook.save("Report_QualityCoding.xlsx")
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
            else:

                if t in task_id_left:
                    old_supp_data = True
                    print("Old task , task id: ", t)
                    if old_supp_data:
                        code_entered0_xpath = "//tr[@class=\"choice_tr saved_data  custom-disabled\"]"
                        if check_exists_by_xpath(driver, code_entered0_xpath):
                            choicevalues0 = extractchoiceinfo_old()
                            codevalues0 = extractcodeinfo_old()
                            dates0 = extractdateinfo_old()
                            ren_provider0 = extractrenderingproviderinfo_old()
                            # print(t, " task : entered ")
                            print("Choice : ", choicevalues0)
                            print("Code : ", codevalues0)
                            print("dates : ", dates0)
                            print("Rendering Provider : ", ren_provider0)
                            try:
                                codestring = code_to_string(t, measure_abbreviation, dates0, choicevalues0, codevalues0,
                                                            ren_provider0)
                            except:
                                err = True
                                codestring = "Blank"

                            if err:
                                sh1.append((str(t), str(choicevalues0), str(codevalues0), str(dates0),
                                            str(ren_provider0), str(driver.current_url)))
                                workbook.save("Report_QualityCoding.xlsx")
                                print("Error : Blank")
                            else:
                                print(codestring)
                            driver.close()
                            driver.switch_to.window(driver.window_handles[0])

                        else:
                            print("Error : No entry in coding sheet ")
                            get_url = driver.current_url
                            sh1.append((str(t), str(choicevalues0), str(codevalues0), str(dates0), str(ren_provider0),
                                        str(driver.current_url)))
                            workbook.save("Report_QualityCoding.xlsx")
                            print(t, get_url)
                            driver.close()
                            driver.switch_to.window(driver.window_handles[0])

                else:
                    print(t, "Error : Task Not found")
                    sh1.append((str(t),"Task Not found"))
                    workbook.save("Report_QualityCoding.xlsx")
                    get_url = driver.current_url
                    print(t, get_url)
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
        next_xpath = "//button[@id='chart_chase_next']"
        next = driver.find_element_by_xpath(next_xpath)
        driver.execute_script("arguments[0].scrollIntoView();", next)
        next.click()

#create Folder or working directory
dateandtime = date_time()
master_directory=config.get("Quality","report_directory_input")
os.chdir(master_directory)
path = makedir(dateandtime)
LOG_FORMAT = "%(levelname)s %(asctime)s - %(message)s"
logging.basicConfig(filename=path + "\\" + "CodingTool-Log.log", level=logging.INFO, format=LOG_FORMAT, filemode='w')
logger = logging.getLogger()
#logger.setLevel(logging.INFO)
os.chdir(path)

downloaddefault=config.get("runner","downloaddefault")
makedir(downloaddefault)
driver = setup("Chrome")
begin_time = datetime.now()
loc = config.get("runner","login_file")

#login
login(driver)
logger.info("Login successful")

#Initialize Worksheet

wb=openpyxl.Workbook()

sm_customer_id = "1300"  # enter customer_id
session_var = 'app_id=smart_chart&custId=' + str(sm_customer_id) + '&payerId=' + str(
    sm_customer_id) + '&orgId=' + str(sm_customer_id)
encoded_string = base64.b64encode(session_var.encode('utf-8'))
driver.get("https://cert.cozeva.com/smart_chart?session=" + encoded_string.decode('utf-8'))

verify_codingsheetQuality(driver, wb, logger, "CozevaSupport")



    # convert into string
    # :26613`MeasureName`Servicedate`CodeValue`RenderingProvider`
    # chalk out a string for BP
end = time.time()
print(f"Total runtime of the program is {end - begin}")
# sys.stdout.close()