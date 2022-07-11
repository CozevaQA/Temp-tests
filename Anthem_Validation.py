from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#defining constants
Measure_abbr_row = 4
counts_start_row = 6
group1_column = 'A'
group5_column = 'B'


path = "C:\\Psuedo D Drive\\AnthemValidation"

wb = load_workbook(path+"\\AnthemCounts.xlsx")
ws = wb.active
wb1 = Workbook()

#what
abbr_row = ws[Measure_abbr_row]
abbr_row2 = ws[3]
length = len(abbr_row)

def generate_sheet(row_counter,sheet_name):
    group1 = ws[group1_column+str(row_counter)]
    group5 = ws[group5_column+str(row_counter)]
    wb1.create_sheet(str(group1.value)+" "+str(group5.value))
    #wb1.create_sheet(sheet_name)
    return str(group1.value)+" "+str(group5.value)



def fill_current_sheet(ws1,current_row,sheet_name):
    ws1.append([sheet_name])
    ws1.append(["Measure Abbr","Numerator count","Denominator Count","Performance"])
    i = 2
    while i < length:
        ws1.append([
            str(abbr_row[i].value)+" "+str(abbr_row2[i].value),str(current_row[i + 1].value),str(current_row[i].value),str(current_row[i + 2].value)])
        i = i + 3


#actual math stuff
row_counter = counts_start_row
nextgroup1 = "x"
x=0
while row_counter<117:
    sheet_name = generate_sheet(row_counter,str(x))
    print(sheet_name)
    current_row = ws[row_counter]
    ws1 = wb1[sheet_name]
    fill_current_sheet(ws1,current_row,sheet_name)
    row_counter = row_counter+1
    x=x+1
    wb1.save(path + "\\AnthemGroupFinal.xlsx")

    #nextgroup1 = ws[group1_column + str(row_counter)].value

wb1.save(path+"\\AnthemGroupFinal.xlsx")





