import timeit

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
#make patient list configurable
# filter in sticket log
import traceback
import base64
import configparser
import sys
from csv import DictReader
import json
import xlrd
from termcolor import colored
from openpyxl.styles.differential import DifferentialStyle
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, \
    ElementNotInteractableException, ElementClickInterceptedException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import xlwt
#from xlwt import Workbook
import time
import csv
from csv import DictReader
from colorama import Fore, Back, Style
import logging
import os
import shutil
import datetime as x
from datetime import date, datetime,timedelta
import pytz
import re
config = configparser.RawConfigParser()
config.read("locator-config.properties")


def setup(val,downloaddefault):
    if val.lower() == "chrome":
        options = Options()
        options.add_argument("--disable-notifications")
        # change for user
        # userdatadir=config.get("runner","user-data-dir")
        # options.add_argument("user-data-dir={}".format(userdatadir))
        # # options.add_experimental_option("detach", True)
        # self.driver = webdriver.Chrome(executable_path="../chromedriver.exe", chrome_options=options)
        # change for user
        preferences = {
            "download.default_directory":downloaddefault}
        options.add_experimental_option("prefs", preferences)
        # change for user
        driverpath=config.get("runner","chromedriverpath")
        driver = webdriver.Chrome(driverpath,options=options)
        driver.get(config.get("runner","Logout_URL"))
        title = driver.get(config.get("runner","Logout_URL"))
        driver.maximize_window()
    return driver


def action_click(element):
    try:
        element.click()
    except (ElementNotInteractableException, ElementClickInterceptedException):
        driver.execute_script("arguments[0].click();", element)

def login(driver, loc):
    # loc = "C:\\Users\\ssrivastava\\Documents\\credentials.xls"
    workbook = xlrd.open_workbook(loc)
    sheet = workbook.sheet_by_index(0)
    uname = driver.find_element_by_id("edit-name")
    pwd = driver.find_element_by_id("edit-pass")
    uname.send_keys(sheet.cell_value(1, 0))
    pwd.send_keys(sheet.cell_value(1, 1))
    driver.find_element_by_id("edit-submit").click()
    # reason for login
    actions = ActionChains(driver)
    WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.ID, "reason_textbox")))
    reason = driver.find_element_by_xpath("//textarea[@id=\"reason_textbox\"]")
    actions.click(reason)
    actions.send_keys_to_element(reason, "https://redmine2.cozeva.com/issues/7662 ")
    actions.perform()
    driver.find_element_by_id("edit-submit").click()


def makedir(foldername):
    path1 = str(foldername)
    if not os.path.exists(path1):
        try:
            os.mkdir(path1)
            return path1
        except OSError as error:
            print(error)
            return False
    else:
        try:
            shutil.rmtree(path1)
            os.mkdir(path1)
            return path1
        except OSError as error:
            print(error)
            return False

def ajax_preloader_wait(driver):
    #time.sleep(1)
    WebDriverWait(driver, 300).until(
        EC.invisibility_of_element((By.XPATH, "//div[contains(@class,'ajax_preloader')]")))
    time.sleep(1)

def date_time():
    today = date.today()
    # print("Today's date:", today)
    tz_In = pytz.timezone('Asia/Kolkata')
    datetime_In = datetime.now(tz_In)
    # print("IN time:", datetime_In.strftime("%I;%M;%S %p"))
    time = datetime_In.strftime("[%I-%M-%S %p]")
    now = str(today) + time
    print(now)
    # logger.info("Date and Time captured!")
    return (now)


def apply_conditional_formatting(sh1):
    red_text = Font(color="9C0006")
    red_fill = PatternFill(bgColor="FFC7CE")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="FAIL", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']

    green_text = Font(color="00FF00FF")
    green_fill = PatternFill(bgColor="0000FF00")
    dxf = DifferentialStyle(font=green_text, fill=green_fill)
    rule1 = Rule(type="containsText", operator="containsText", text="PASS", dxf=dxf)
    rule1.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']

    sh1.conditional_formatting.add('B1:B10000', rule)
    sh1.conditional_formatting.add('B1:B10000', rule1)

#create Folder or working directory
# dateandtime = date_time()
# master_directory=config.get("runner","report_directory_input_markaspending")
# #can include folder path here
# os.chdir(master_directory)
# path = makedir(dateandtime)
# LOG_FORMAT = "%(levelname)s %(asctime)s - %(message)s"
# logging.basicConfig(filename=path + "\\" + "Mark-As-Pending-Log.log", level=logging.INFO, format=LOG_FORMAT, filemode='w')
# logger = logging.getLogger()
# #logger.setLevel(logging.INFO)
# os.chdir(path)

# downloaddefault=config.get("runner","downloaddefault")
# makedir(downloaddefault)
# driver = setup("Chrome",downloaddefault)
# begin_time = datetime.now()
# loc = config.get("runner","login_file")
#
# #login
# login(driver, loc)
#logger.info("Login successful")




def open_registry_page(customer_id):
    customer_list_url = []
    sm_customer_id = str(customer_id)
    sm_customer_id = sm_customer_id.split(".")[0]
    session_var = 'app_id=registries&custId=' + str(sm_customer_id) + '&payerId=' + str(
        sm_customer_id) + '&orgId=' + str(sm_customer_id)
    encoded_string = base64.b64encode(session_var.encode('utf-8'))
    customer_list_url.append(encoded_string)
    for idx, val in enumerate(customer_list_url):
        driver.get(config.get("runner","URL")+"registries?session=" + val.decode('utf-8'))

def check_exists_by_xpath(driver,xpath):
    try:
        driver.find_element_by_xpath(xpath)
        print("element exists")
        return True
    except NoSuchElementException:
        return False


def PatientDashboard(driver, sheet, quarter_name, lob_name, customer_id, metric_name_4_patientdashboard,
                     add_supdata_flag_MSPL,
                     map_flag_MSPL, caregap_MSPL, mspl_url, provider_name):
    # driver.get("https://stage.cozeva.com/patient_detail/1R0ADY3?tab_type=CareOps&cozeva_id=1R0ADY3&patient_id=9290597&cozeva_id=1R0ADY3&session"
    #          "=YXBwX2lkPXJlZ2lzdHJpZXMmY3VzdElkPTE1MDAmZG9jdG9yc1BlcnNvbklkPTExODUxNTYzJmRvY3Rvcl91aWQ9MTE4MzE0ODkmcGF5ZXJJZD0xNTAwJnF1YXJ0ZXI9MjAyMC0xMi0zMSZob21lPVlYQndYMmxrUFhKbFoybHpkSEpwWlhNbVkzVnpkRWxrUFRFMU1EQW1jR0Y1WlhKSlpEMHhOVEF3Sm05eVowbGtQVEUxTURB&first_load=1")
    global add_supdata_flag_pt, map_flag_pt, cozeva_id, pcp_name


    try:
        ajax_preloader_wait(driver)
        if len(driver.find_elements_by_xpath("(//div/span[@data-tooltip='Cozeva Id (Click to Copy)'])[1]")) != 0:
            cozeva_id = driver.find_element_by_xpath("(//div/span[@data-tooltip='Cozeva Id (Click to Copy)'])[1]").text
        elif len(driver.find_elements_by_xpath("(//div/span[@data-tooltip='Cozeva Id (Click to Copy)'])[1]")) == 0:
            cozeva_id = "Blank ; Please check "
            return
        sheet['B2'] = quarter_name + " | " + lob_name
        sheet['B3'] = metric_name_4_patientdashboard
        sheet['B4'] = cozeva_id
        # find metric pencil icon

        # Red dot count
        caregap_pt = len(driver.find_elements_by_xpath("//div[@class='non_compliant red_dot']"))

        metric_name = metric_name_4_patientdashboard
        metrics_patientdashboard = driver.find_elements_by_xpath("//div[@class='text-bold sub-title']")
        measure_display_flag = 0
        for metric_counter in range(len(metrics_patientdashboard)):
            print("Metric Counter" + str(metric_counter))
            metric_counter1 = metric_counter + 1
            xpath1 = "(" + "//div[@class='text-bold sub-title']" + ")" + "[" + str(metric_counter1) + "]"
            xpath_metric_row = xpath1 + "/../../../../../.."
            xpath_pencil_patientdashboard = xpath1 + "/../../../../../../td/div/div[@class='dropdown']//child::a[@class='addSuppData-trigger pts']//child::i"
            metric_name_patientdashboard = metrics_patientdashboard[metric_counter].text
            print("Metric name in patient dashboard" + str(metric_name_patientdashboard))
            metric_row = driver.find_element_by_xpath(xpath_metric_row)
            # print(metric_name_patientdashboard)#Print all measures in Patient dashboard

            if metric_name_patientdashboard == metric_name:

                ActionChains(driver).move_to_element(metric_row).perform()

                if len(driver.find_elements_by_xpath(xpath_pencil_patientdashboard)) == 1:
                    sheet.append(("Pencil icon Present ?", "PASS"))
                    action_click(driver.find_element_by_xpath(xpath_pencil_patientdashboard))
                    print("Clicked on Pencil icon")
                    time.sleep(1)
                    xpath_pencil_options = xpath_pencil_patientdashboard + "//..//..//child::ul/li"
                    WebDriverWait(driver, 30).until(
                        EC.visibility_of_element_located((By.XPATH, xpath_pencil_options)))
                    pencil_options = driver.find_elements_by_xpath(xpath_pencil_options)
                    add_supdata_flag_pt = 0
                    map_flag_pt = 0

                    for option_counter in range(len(pencil_options)):

                        print((pencil_options[option_counter]).text)
                        pencil_options_pt_text = (pencil_options[option_counter]).text
                        if pencil_options_pt_text.strip() == "Add Supplemental Data":
                            add_supdata_flag_pt = 1
                            sheet.append(("Add Supplemental Data Present ?", "PASS"))
                            action_click(pencil_options[option_counter])
                            ajax_preloader_wait(driver)
                            # verify submit button
                            submit_button_xpath = config.get("MAP", "submit_xpath")
                            try:
                                driver.find_element_by_xpath(submit_button_xpath)
                                sheet.append(("Submit button appearing in Supp data", "PASS"))
                            except NoSuchElementException:
                                sheet.append(("Submit button appearing in Supp data", "FAIL"))

                            # verify delete button
                            delete_button_xpath = config.get("MAP", "delete_xpath")
                            try:
                                driver.find_element_by_xpath(delete_button_xpath)
                                sheet.append(("Delete button appearing in Supp data", "PASS"))
                            except NoSuchElementException:
                                sheet.append(("Delete button appearing in Supp data", "FAIL"))

                            # verify Task id
                            task_text_xpath = config.get("MAP", "task_id_xpath")
                            try:
                                task_text = driver.find_element_by_xpath(task_text_xpath).text
                                sheet.append(("Task id appearing in Supp data", "PASS", str(task_text)))
                            except NoSuchElementException:
                                sheet.append(("Task id appearing in Supp data", "FAIL"))

                            # verify attachment section
                            attachment_xpath = config.get("MAP", "attachment_xpath")
                            try:
                                driver.find_element_by_xpath(attachment_xpath)
                                sheet.append(("Attachment section appearing in Supp data", "PASS"))
                            except NoSuchElementException:
                                sheet.append(("Attachment section appearing in Supp data", "FAIL"))

                            # Delete the task
                            delete_button_xpath = config.get("MAP", "delete_xpath")
                            try:
                                action_click(driver.find_element_by_xpath(delete_button_xpath))

                                # give reason
                                reason_modal_xpath = config.get("MAP", "reason_input_modal")
                                reason_modal = driver.find_element_by_xpath(reason_modal_xpath)
                                reason_modal.send_keys("Cozeva QA")
                                action_click(driver.find_element_by_xpath(config.get("MAP", "confirm_modal_xpath")))
                                time.sleep(5)
                                ajax_preloader_wait(driver)

                                sheet.append(("Task Deleted", "PASS"))
                            except NoSuchElementException:
                                sheet.append(("Task Deleted", "FAIL", "Manual intervention required "))

                            action_click(driver.find_element_by_xpath(xpath_pencil_patientdashboard))

                        if pencil_options_pt_text.strip() == "Mark as Pending":
                            sheet.append(("Mark As Pending Present ?", "PASS"))
                            # click on MAP
                            map_flag_pt = 1
                            action_click(pencil_options[option_counter])

                            # click on confirm
                            action_click(driver.find_element_by_xpath(config.get("MAP", "confirm_modal_xpath")))

                            # wait for page to load
                            ajax_preloader_wait(driver)

                            # check for stale icon
                            restored = 0
                            stale_icon = 0
                            x = 1
                            start_time1 = time.time()
                            while True:
                                driver.refresh()
                                ajax_preloader_wait(driver)
                                if (check_exists_by_xpath(driver, config.get("MAP", "stale_icon_xpath"))):
                                    print("Stale icon found ")
                                    stale_icon = 1
                                    break
                                if (x == 10):
                                    break
                                x = x + 1
                            time_elapsed1_value = timeit.default_timer() - start_time1
                            time_elapsed1 = '{0:.2f}'.format(time_elapsed1_value)
                            if (stale_icon == 1):
                                timestring = "Time taken " + str(time_elapsed1)
                                sheet.append(("Mark As pending - Stale icon ", "PASS", str(timestring)))
                            else:
                                timestring = "Time taken " + str(time_elapsed1)
                                sheet.append(("Mark As pending -Stale icon", "FAIL", str(timestring)))

                            # Keep refreshing till you see the hollow dot
                            # Refresh 10 times to verify appearing of hollow dot
                            start_time = timeit.default_timer()
                            hollow_dot_found = 0
                            dot_status_xpath = "(//div[@class='text-bold sub-title'])" + "[" + str(
                                metric_counter1) + "]" + "//ancestor::tr//child::td[1]//child::div[contains(@style,'margin: 8px 0px 0px 4px;')]"
                            dot_status = driver.find_element_by_xpath(dot_status_xpath).get_attribute("class")
                            print("Dot status " + str(dot_status))
                            y = 1
                            while True:
                                driver.refresh()
                                ajax_preloader_wait(driver)
                                dot_status_xpath = "(//div[@class='text-bold sub-title'])" + "[" + str(
                                    metric_counter1) + "]" + "//ancestor::tr//child::td[1]//child::div[contains(@style,'margin: 8px 0px 0px 4px;')]"
                                dot_status = driver.find_element_by_xpath(dot_status_xpath).get_attribute("class")
                                if (dot_status == "non_compliant hollow_dot"):
                                    print("Checking for hollow dot ")
                                    hollow_dot_found = 1
                                    break
                                if (y == 10):
                                    break
                                print("Dot Status while checking for hollow dot " + str(dot_status))
                                y = y + 1

                            elapsed_value = timeit.default_timer() - start_time
                            elapsed='{0:.2f}'.format(elapsed_value)
                            if (hollow_dot_found == 1):  # Click on pencil icon and unmark as Pending
                                ajax_preloader_wait(driver)
                                timestring = "Time taken " + str(elapsed)
                                sheet.append(("Hollow dot ", "PASS ", str(timestring)))
                                WebDriverWait(driver, 30).until(
                                    EC.element_to_be_clickable((By.XPATH, xpath_pencil_patientdashboard)))
                                action_click(driver.find_element_by_xpath(xpath_pencil_patientdashboard))
                                print("Clicked on Pencil icon")
                                unmark_as_pending_xpath = '(//*[text()="Unmark as Pending"])[' + str(
                                    metric_counter1) + ']'
                                unmark_as_pending = driver.find_element_by_xpath(unmark_as_pending_xpath)
                                action_click(unmark_as_pending)
                                print("Clicked on unmark as pending icon")
                                ajax_preloader_wait(driver)
                                dot_status_xpath = "(//div[@class='text-bold sub-title'])" + "[" + str(
                                    metric_counter1) + "]" + "//ancestor::tr//child::td[1]//child::div[contains(@style,'margin: 8px 0px 0px 4px;')]"
                                dot_status = driver.find_element_by_xpath(dot_status_xpath).get_attribute("class")
                                z = 0
                                while True:
                                    driver.refresh()
                                    ajax_preloader_wait(driver)
                                    dot_status_xpath = "(//div[@class='text-bold sub-title'])" + "[" + str(
                                        metric_counter1) + "]" + "//ancestor::tr//child::td[1]//child::div[contains(@style,'margin: 8px 0px 0px 4px;')]"
                                    dot_status = driver.find_element_by_xpath(dot_status_xpath).get_attribute("class")
                                    if (dot_status == "non_compliant red_dot"):
                                        restored = 1
                                        break
                                    if (z == 10):
                                        break
                                    z = z + 1
                                if (restored == 1):
                                    timestring = "Time waited for hollow dot " + str(elapsed)
                                    sheet.append(
                                        ("Unmark as Pending", "PASS", str(timestring)))
                                else:
                                    timestring = "Manual intervention required , Time waited for hollow dot " + str(elapsed)
                                    sheet.append(
                                        ("Unmark as pending hasn't occurred as red dot has not re-appear", "FAIL",
                                         str(timestring)))
                            else:
                                timestring = "Manual intervention required , Time waited for hollow dot " + str(elapsed)
                                sheet.append(("Unmark as pending hasn't occurred as hollow dot has not appear", "FAIL",
                                              str(timestring)))

                        pencil_options = driver.find_elements_by_xpath(xpath_pencil_options)
                    if (add_supdata_flag_pt != 1):
                        sheet.append(("Add Supplemental Data Present ?", "FAIL"))
                    if (map_flag_pt != 1):
                        sheet.append(("Mark As Pending Present ?", "FAIL"))



                elif len(driver.find_elements_by_xpath(xpath_pencil_patientdashboard)) == 0:
                    print("NO PENCIL")
                    sheet.append(("Pencil icon Present ?", "FAIL"))
                    add_supdata_flag_pt = 0
                    map_flag_pt = 0
                    return False

                print("Supdata flag(Pt): " + str(add_supdata_flag_pt))
                print("Map flag(Pt): " + str(map_flag_pt))
                measure_display_flag = 1
                break
            else:
                print("Metric name is not equal")
                return False
        if (add_supdata_flag_pt == map_flag_pt == 1):
            return True
        else:
            return False

    except Exception as e:
        print(e)
        return False

        #Click on MAP
        #confirm yes on the modal
        #check stale icon
        #4 th test case pass
        #click on Add Supp data option
        #verify submit and delete button
        #5th Test Case pass
        #navigate to pending list
        #check for patient cozeva id
        #refresh 5-6 times
        #6h test case pass
        #return to Patient dashboard
        #check for hollow dot
        #unmark as pending
        #7th test case pass
        #check for no hollow dot


def verify_mark_as_pending(driver, workbook, logger, run_from, customer_id):
    ws1 = workbook.create_sheet("MAPCodingTool")
    sh1 = workbook["MAPCodingTool"]
    sh1['A1'].value = "Test Data"
    sh1['A1'].font = Font(bold=True, size=13)
    sh1['A2'].value = "LOB"
    sh1['A2'].font = Font(bold=True, size=13)
    sh1['A3'].value = "Metric"
    sh1['A3'].font = Font(bold=True, size=13)
    sh1['A4'].value = "Cozeva ID"
    sh1['A4'].font = Font(bold=True, size=13)
    sh1['A5'].value="Test Case"
    sh1['A5'].font = Font(bold=True, size=13)
    sh1['B5'].value = "Status"
    sh1['B5'].font = Font(bold=True, size=13)
    sh1['C5'].value = "Comments"
    sh1['C5'].font = Font(bold=True, size=13)
    patient_verified=""
    # navigate to registry
    open_registry_page(customer_id)
    ajax_preloader_wait(driver)
    #customer_name = driver.find_element_by_xpath(config['LOCATOR']['xpath_contextName']).text
    driver.find_element_by_xpath("//a[@id='qt-filter-label']").click()
    time.sleep(1)
    quarters = driver.find_elements_by_xpath("//ul[@id='filter-quarter']/li")
    lobs = driver.find_elements_by_xpath("//ul[@id='filter-lob']/li[@class!='hide']")
    driver.find_element_by_xpath("//a[@id='qt-filter-label']").click()
    patient_found = ""
    for quarter in range(2):
        # quarter = quarter + 1
        if (patient_found == "Found"):
            break
        for lob in range(len(lobs)):
            if (patient_found == "Found"):
                break
            # for lob in range(1):
            # lob = lob + 3
            time.sleep(0.5)
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//a[@id='qt-filter-label']")))
            driver.find_element_by_xpath("//a[@id='qt-filter-label']").click()
            time.sleep(0.25)
            quarter_name = quarters[quarter].text
            print(colored(quarter_name, 'blue'))
            quarters[quarter].click()
            time.sleep(0.25)
            lobs[lob].click()
            lob_name = lobs[lob].text
            print(colored(lob_name, 'magenta'))
            driver.find_element_by_xpath("//a[@id='reg-filter-apply']").click()
            WebDriverWait(driver, 90).until(EC.invisibility_of_element((By.XPATH, "//div[@class='ajax_preloader']")))
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, "//a[@data-target='qt-reg-nav-filters']")))
            driver.find_element_by_xpath("//a[@data-target='qt-reg-nav-filters']").click()
            time.sleep(0.25)
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//label[@class='col s12']")))
            driver.find_element_by_xpath("//label[@class='col s12']").click()
            time.sleep(0.25)
            driver.find_element_by_xpath("//button[@id='qt-apply-search']").click()
            WebDriverWait(driver, 90).until(EC.invisibility_of_element((By.XPATH, "//div[@class='ajax_preloader']")))

            """
            **** SUPPORT MEASURE REGISTRY NAVIGATION ****
            """
            measures_all = driver.find_elements_by_xpath("//div/span[@class='met-name']")
            scores = driver.find_elements_by_xpath("//span[@class='num-den']")
            measure_counter = 0
            score = 0

            while measure_counter < len(measures_all) and score < len(scores):
                if (patient_found == "Found" or patient_verified==True):
                    break
                WebDriverWait(driver, 60).until(
                    EC.presence_of_element_located((By.XPATH, "//a[@id='reg-faq-trigger']")))
                time.sleep(0.5)
                driver.execute_script("arguments[0].scrollIntoView();", measures_all[measure_counter])
                measure_name = (measures_all[measure_counter]).text
                print("Measure name: ", measure_name)
                numdeno = scores[score].text
                numdeno = numdeno.lstrip("(")
                numdeno = numdeno.rstrip(")")
                numdeno = numdeno.split("/")
                Numerator = numdeno[0].replace(',', '')
                Denominator = numdeno[1].replace(',', '')
                print("Numerator=", Numerator)
                print("Denominator=", Denominator)
                last_url = driver.current_url
                measures_all[measure_counter].click()
                try:
                    ajax_preloader_wait(driver)
                    if len(driver.find_elements_by_xpath("//td/a[contains(@href,'/registries/')]")) == 0 and float(
                            Denominator) != 0 and float(Numerator) != 0:
                        met_name = driver.find_element_by_xpath(
                            "//div[@class='ch metric_specific_patient_list_title']").text
                        logger.info("Metric name: %s", measure_name)
                        logger.warning("Providers list is blank. Please check manually.")
                        sh1.append((quarter_name + " | " + lob_name, measure_name, 'Blank Providers List'))

                    elif len(driver.find_elements_by_xpath("//td/a[contains(@href,'/registries/')]")) == 0 and float(
                            Denominator) == 0:
                        met_name = driver.find_element_by_xpath(
                            "//div[@class='ch metric_specific_patient_list_title']").text
                        logger.info("Metric name: %s", measure_name)
                        logger.info("Providers list is blank since measure score is zero.")

                    elif len(driver.find_elements_by_xpath("//td/a[contains(@href,'/registries/')]")) == 0:
                        met_name = driver.find_element_by_xpath(
                            "//div[@class='ch metric_specific_patient_list_title']").text
                        logger.info("Metric name: %s", measure_name)
                        logger.info("Providers list is blank. Please check manually.")

                    else:
                        if len(driver.find_elements_by_xpath("(//td/a[contains(@href,'/registries/')])[2]")) != 0:
                            patientlist_link = driver.find_element_by_xpath(
                                "(//td/a[contains(@href,'/registries/')])[2]")

                        else:
                            patientlist_link = driver.find_element_by_xpath(
                                "(//td/a[contains(@href,'/registries/')])[1]")

                        ActionChains(driver).move_to_element(patientlist_link).perform()
                        ActionChains(driver).key_down(Keys.CONTROL).click(patientlist_link).key_up(
                            Keys.CONTROL).perform()

                        # """ *********** Termed column check **************** """
                        # ajax_preloader_wait()
                        # if len(driver.find_elements_by_xpath("//th[@aria-label='Termed: activate to sort column ascending']"))!=0:
                        #     print("Termed column exists!")
                        # else:
                        #     print("Please check")
                        # driver.find_element_by_xpath("//a[@class='datatable_filter_dropdown sidenav-trigger']").click()
                        # time.sleep(1)
                        #
                        # if len(driver.find_elements_by_xpath("//div[text()='Termed:']"))!=0:
                        #     print("Filter is present")
                        # else:
                        #     print("Check Filter")

                        # **** CALCULATE CARE GAP LIST ****
                        try:
                            driver.switch_to.window(driver.window_handles[1])
                            ajax_preloader_wait(driver)

                            provider_name = driver.find_element_by_xpath("//a[@id='context_trigger']/div/span").text
                            logger.info("Provider Name: %s", provider_name)
                            metric_name = driver.find_element_by_xpath(
                                "//div[@class='ch metric_specific_patient_list_title']").text
                            logger.info("Metric name: %s", metric_name)
                            my_lob_ce = driver.find_element_by_xpath(
                                "//div[@class='metric_patient_list_filter left']").text
                            print(my_lob_ce)
                            x = my_lob_ce.split("\u2002·\u2002")
                            my_lob_ce_final = x[0] + " " + "|" + " " + x[1] + " " + "|" + " " + x[3]
                            logger.info("%s", my_lob_ce_final)

                            # Data for patient dashboard:
                            y = metric_name.split("|")

                            metric_name_4_patientdashboard1 = y[1].strip()
                            metric_name_4_patientdashboard = metric_name_4_patientdashboard1.replace('*', '')
                            print(metric_name_4_patientdashboard)

                            if len(driver.find_elements_by_xpath(
                                    "//td/div/a[contains(@href,'/patient_detail/')]")) == 0:
                                print("Patient list is blank!")
                                sh1.append((quarter_name + " | " + lob_name, metric_name_4_patientdashboard,
                                            provider_name, 'No Non-compliant patient found'))
                                measure_name4screenshot = ''.join(
                                    e for e in str(measure_name) if (e.isalnum() or e.isspace()))


                            # Pencil icon presence:
                            elif len(driver.find_elements_by_xpath(
                                    "//td/div/a[contains(@href,'/patient_detail/')]")) != 0:
                                time.sleep(1)
                                if len(driver.find_elements_by_xpath("//td[contains(@class,' pencil_icon')]")) != 0:
                                    driver.find_element_by_xpath("//td[contains(@class,' pencil_icon')]").click()
                                    time.sleep(0.5)
                                    pencil_options = driver.find_elements_by_xpath(
                                        "(//td[contains(@class,' pencil_icon')])[1]/div/ul[contains(@class,'dropdown-content patient-menu-list')]/li")
                                    # Available options in Pencil icon:
                                    add_supdata_flag_MSPL = 0
                                    map_flag_MSPL = 0
                                    option_counter = 0

                                    for option_counter in range(len(pencil_options)):
                                        print((pencil_options[option_counter]).text)
                                        pencil_options_text = (pencil_options[option_counter]).text
                                        if pencil_options_text.strip() == "Add Supplemental Data":
                                            add_supdata_flag_MSPL = 1
                                        elif pencil_options_text.strip() == "Mark as Pending":
                                            map_flag_MSPL = 1
                                        elif pencil_options_text.strip() == "Confirm/Disconfirm":
                                            add_supdata_flag_MSPL = "Confirm/Disconfirm"
                                        pencil_options = driver.find_elements_by_xpath(
                                            "(//td[contains(@class,' pencil_icon')])[1]/div/ul[contains(@class,'dropdown-content patient-menu-list')]/li")
                                        if(map_flag_MSPL==add_supdata_flag_MSPL==1):
                                            patient_found="Found"

                                # Pencil icon is not present:
                                elif len(driver.find_elements_by_xpath("//td[contains(@class,' pencil_icon')]")) == 0:
                                    print("No Pencil in MSPL")
                                    add_supdata_flag_MSPL = 0
                                    map_flag_MSPL = 0
                                print("Supdata flag(MSPL): " + str(add_supdata_flag_MSPL))
                                print("Map flag(MSPL): " + str(map_flag_MSPL))



                                # CareGap in MSPL:
                                if len(driver.find_elements_by_xpath("//td[contains(@class,'care_ops')]")) != 0:
                                    caregap_MSPL = driver.find_element_by_xpath(
                                        "(//td[contains(@class,'care_ops')])[1]").text
                                    print("CareGap in MSPL:" + caregap_MSPL)
                                elif len(driver.find_elements_by_xpath("//td[contains(@class,' care_ops')]")) == 0:
                                    caregap_MSPL = "Not present"
                                    print("MSPL: CareGap is Not present")

                                # call PATIENT DASHBOARD:
                                mspl_url = driver.current_url
                                driver.find_element_by_xpath("//td/div/a[contains(@href,'/patient_detail/')]").click()
                                try:
                                    driver.switch_to.window(driver.window_handles[2])

                                    if(map_flag_MSPL==1 and add_supdata_flag_MSPL==1):
                                        patient_verified = PatientDashboard(driver,sh1,quarter_name, lob_name, customer_id,
                                                         metric_name_4_patientdashboard, add_supdata_flag_MSPL,
                                                         map_flag_MSPL, caregap_MSPL, mspl_url,provider_name)



                                except Exception as e:
                                    print(e)
                                finally:

                                    driver.close()
                                    driver.switch_to.window(driver.window_handles[1])





                        # Exception in MSPL block
                        except Exception as e:
                            print(e)
                            logger.critical(
                                measure_name + '\n' + provider_name + '\n' + "Metric specific patients list is not opening!Exception occurred!!")
                            # sh1.append((quarter_name + " | " + lob_name, measure_name,
                            #             provider_name, 'Error'))

                        finally:
                            driver.close()
                            driver.switch_to.window(driver.window_handles[0])
                            apply_conditional_formatting(sh1)
                            wb.save( "Report.xlsx")

                    WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, "//a[@class='breadcrumb']")))
                    driver.find_element_by_xpath("//a[@class='breadcrumb']").click()

                # Providers list open exception block
                except Exception as e:

                    print(e)
                    driver.get(last_url)
                finally:
                    # MEASURE COUNTER

                    measures_all = driver.find_elements_by_xpath("//div/span[@class='met-name']")
                    scores = driver.find_elements_by_xpath("//span[@class='num-den']")
                    measure_counter += 1
                    score += 1

            lobs = driver.find_elements_by_xpath("//ul[@id='filter-lob']/li[@class!='hide']")
            quarters = driver.find_elements_by_xpath("//ul[@id='filter-quarter']/li")


#store Cozeva ID
#search the metric
#Check options in Pencil icon
#Click on Mark As Pending
#Page will refresh ;Check for stale icon
#Click on Add Supplemental Data
#Check Task Id
#Check reflection on Pending List  compare patient and metric name
#Refresh twice or thrice
#Come to patient dashboard
#Unmark As pending


#initialize Workbook

verify_mark_as_pending(driver,wb,logger,"CS",customer_id)