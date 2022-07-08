from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

path = "C:\\Psuedo D Drive\\AnthemValidation"

wb = load_workbook(path+"\\AnthemCounts.xlsx")